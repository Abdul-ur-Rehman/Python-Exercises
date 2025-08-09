import openpyxl as excel

book = excel.load_workbook(r"D:\QA Automation\PythonExercises\test_excel.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)

print(sheet.max_row)
print(sheet.max_column)

#sheet.cell(row=2, column=2).value = "Abdur Rehman"
print(sheet.cell(row=2, column=2).value)
Dict = {}
list = []

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "test_HomePage_FillingForm":
        for j in range(2, sheet.max_column + 1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        list.append(Dict)

print(list)